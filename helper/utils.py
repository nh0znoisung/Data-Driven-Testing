import undetected_chromedriver as uc
import sys
import os
sys.path.append(os.path.abspath('../helper'))
from helper.Login import Login
from helper.ChangePassword import ChangePassword
from dotenv import load_dotenv
load_dotenv()

import openpyxl
import pytest



def login_func(username: str, password: str) -> bool:
    """Login function for testing"""
    login_ins = Login(uc.Chrome(use_subprocess=True))
    login_res = login_ins.login_run(username=username, password=password)
    login_ins.stop()
    return login_res


def changepass_func(current_password: str, new_password: str, new_password_again: str) -> tuple[bool,tuple[str,str,str]]: 
    """Change password function for testing"""
    changepassword_ins = ChangePassword(uc.Chrome(use_subprocess=True))
    changepwd_res = changepassword_ins.changepwd_run(current_password, new_password, new_password_again)
    changepassword_ins.stop()
    return changepwd_res

def load_excel(filename: str='excel.xlsx', sheetname: str='Sheet1') -> list[list[str]]:
    """Load the whole data from excel file and sheet name"""
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.get_sheet_by_name(sheetname)
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    # print(totalrows, totalcols)

    # The first line is the header
    data_list = [ [sheet.cell(i,j).value if sheet.cell(i,j).value is not None else ''  for j in range(1, totalcols+1)] for i in range(2, totalrows+1) ]
    return data_list

B_F1_title = '_username,_password,_expect'
B_F2_title = '_current_password,_new_password,_new_password_again,_expect_res,_expect_currentpwd_msg,_expect_newpwd_msg,_expect_newpwdagain_msg'